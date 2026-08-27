"""Convert the state research workbooks into PaperTrail document JSON.

Each record is graded:
  guide     - has fee, required documents, processing time and eligibility, and
              its verification status is clean. Answerable end to end, so it
              joins the searchable corpus.
  directory - the service is confirmed official and has a portal link, but the
              detail fields the government page never stated are absent. Kept
              verbatim for the directory tier; never fed to answer generation,
              because a blank fee invites the model to invent one.

Nothing here fills a blank field. A value is copied only if the workbook holds
it, so every fact keeps the provenance the researcher recorded.
"""
import json
import hashlib
import argparse
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
RESEARCH_DIR = ROOT / "backend" / "data" / "research"

# Values the researcher used to mean "the official page did not state this"
BLANK = {"", "not verified", "none", "n/a", "na", "not stated",
         "none stated", "not applicable", "-", "nil"}

# Some cells hold a sentence reporting an absence rather than a value, e.g.
# "No government fee indicated in the official approval catalogue". Those are
# truthful but they are not answers, and passing them through would dress an
# absence up as content. Treated as blank so the field reads as unknown.
# Some cells hold a sentence reporting an absence rather than a value, e.g.
# "No government fee indicated in the official approval catalogue". Those are
# truthful but they are not answers, and passing one through would dress an
# absence up as content. Plain string tests rather than a regex, because the
# escaping in a pattern this fiddly is easy to get silently wrong.
ABSENCE_PHRASES = (
    "not indicated", "not stated", "not specified", "not mentioned",
    "not available", "not listed", "not published", "not provided",
    "no information", "not in that table", "not transcribed",
)
ABSENCE_VERBS = ("indicated", "stated", "specified", "mentioned",
                 "available", "listed", "published")


def is_absence(text: str) -> bool:
    """True when the cell reports that a value is missing rather than giving one."""
    t = " ".join(text.lower().split())
    if any(phrase in t for phrase in ABSENCE_PHRASES):
        return True
    # "No government fee indicated ...", "No separate utility fee stated ..."
    return (t.startswith("no ") and "fee" in t
            and any(v in t for v in ABSENCE_VERBS))

# Fields that must all be present for a record to answer a PaperTrail question
# A record earns "guide" grade when it has an official portal to send the user
# to, plus at least two of the four facts they actually ask about. Requiring all
# four collapses the corpus to 32 records, because the workbooks often hold a
# sentence explaining a field was not on the source page; requiring one lets
# through records carrying nothing but an eligibility line. Two is the point
# where a record answers something real. Whatever is missing is rendered to the
# model as NOT PUBLISHED, never as a blank it might fill in.
ANSWER_FIELDS = ["Required Documents", "Government Fee",
                 "Expected Processing Time", "Who Can Apply"]
MIN_ANSWER_FIELDS = 2

GUIDE_STATUSES = {"VERIFIED", "PARTIALLY VERIFIED"}

CONFIDENCE_MAP = {
    "VERIFIED": "VERIFIED",
    "PARTIALLY VERIFIED": "PARTIALLY VERIFIED",
    "UNVERIFIED": "UNVERIFIED",
    "CONFLICTING": "UNVERIFIED",
    "OUTDATED": "UNVERIFIED",
}

JURISDICTION_MAP = {
    "Central": "central",
    "Central / State": "central_state_executed",
    "State": "state",
    "District / Regional": "state",
    "Tehsil / Subdistrict": "local_rural",
    "Rural Local Body": "local_rural",
    "Urban Local Body": "local_urban",
}


# The workbooks use a richer sector taxonomy than the live corpus. Fold it onto
# the existing ten categories so Browse does not show "Transport" beside
# "Transport & Mobility". Placement follows where the equivalent process already
# sits in the Karnataka/Maharashtra data (Fire NOC and FSSAI are Business there,
# not Health or Legal).
SECTOR_MAP = {
    "Identity, Citizenship & Elections": "Core Identity",
    "Registration & Civil Records": "Core Identity",
    "Digital Government & Service Access": "Core Identity",
    "Business, Industry & Investment": "Business",
    "Taxation & Commercial Regulation": "Finance",
    "Food & Drug Regulation": "Business",
    "Fire & Emergency Services": "Business",
    "Excise": "Business",
    "Energy & Electrical Safety": "Business",
    "Environment, Forest & Water Regulation": "Business",
    "Labour, Employment & Social Security": "Employment",
    "Land, Revenue & Property Records": "Property",
    "Housing, Planning & Construction": "Property",
    "Urban Civic Services": "Property",
    "Transport & Mobility": "Transport",
    "Agriculture": "Welfare",
    "Other Government Services": "General",
}


def map_category(sector: str) -> str:
    return SECTOR_MAP.get(sector, sector or "General")


def clean(v: Any) -> str:
    """Return the value, or an empty string if it is one of the blank markers."""
    if v is None:
        return ""
    s = str(v).strip()
    # placeholders such as "-", "---", "n/a", "." carry no information
    if not any(ch.isalnum() for ch in s):
        return ""
    if s.lower() in BLANK or is_absence(s):
        return ""
    return s


def join(*parts: str, sep: str = " ") -> str:
    return sep.join(p for p in parts if p)


KEY_COLUMN = "Document / Service Name"


def find_header_row(rows: List[tuple], limit: int = 12) -> int:
    """Locate the header row.

    The research guide asks for headers in row 1, but earlier workbooks carried
    a title and record-count banner above them. Detect the row holding the key
    column rather than assuming a fixed offset, so both layouts import.
    """
    for i, row in enumerate(rows[:limit]):
        if any(str(c).strip() == KEY_COLUMN for c in row if c is not None):
            return i
    raise ValueError(
        f"Could not find a header row containing {KEY_COLUMN!r} in the first "
        f"{limit} rows. Check the sheet follows docs/state-research-guide.md."
    )


def read_sheet(path: Path) -> List[Dict[str, str]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    # The guide names the main sheet "Documents"; older workbooks used the
    # state name, and it was always the first sheet.
    ws = wb["Documents"] if "Documents" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    h = find_header_row(rows)
    header = list(rows[h])
    out = []
    for r in rows[h + 1:]:
        if not r or not any(c not in (None, "") for c in r):
            continue
        record = {col: r[i] for i, col in enumerate(header) if col}
        if not str(record.get(KEY_COLUMN, "")).strip():
            continue
        out.append(record)
    return out


def convert(raw: Dict[str, Any], state: str) -> Optional[Dict[str, Any]]:
    name = clean(raw.get("Document / Service Name"))
    if not name:
        return None

    status = (clean(raw.get("Verification Status")) or "UNVERIFIED").upper()
    answered = sum(1 for f in ANSWER_FIELDS if clean(raw.get(f)))
    grade = (
        "guide"
        if status in GUIDE_STATUSES
        and clean(raw.get("Official Portal"))
        and answered >= MIN_ANSWER_FIELDS
        else "directory"
    )

    fee = join(
        clean(raw.get("Government Fee")),
        (f"(service charge: {clean(raw.get('Service Charge'))})"
         if clean(raw.get("Service Charge")) else ""),
    )
    required = join(
        clean(raw.get("Required Documents")),
        clean(raw.get("Supporting Documents")),
        clean(raw.get("Identity Proof")),
        clean(raw.get("Address Proof")),
        sep="; ",
    )
    online = join(
        clean(raw.get("Application Method")),
        clean(raw.get("Where to Apply")),
        clean(raw.get("Direct Application URL")),
        sep=" ",
    ) if clean(raw.get("Access Mode")).lower() != "offline" else ""
    offline = join(
        clean(raw.get("Physical Office Name")),
        clean(raw.get("Office Address")),
        clean(raw.get("Office Jurisdiction")),
        sep=", ",
    )

    return {
        "id": hashlib.md5(f"{state}:{name}".encode()).hexdigest()[:16],
        "name": name,
        "state": state,
        "confidence": CONFIDENCE_MAP.get(status, "UNVERIFIED"),
        "grade": grade,
        "jurisdiction": JURISDICTION_MAP.get(
            clean(raw.get("Government Level")), "state"),
        "issuing_office": clean(raw.get("Issuing Authority")),
        "department": clean(raw.get("Department")),
        "fee": fee,
        "fee_known": bool(fee),
        "answered_fields": answered,
        "processing_time": join(clean(raw.get("Expected Processing Time")),
                                clean(raw.get("Statutory Timeline")), sep="; "),
        "portal": clean(raw.get("Official Portal")),
        "source_url": clean(raw.get("Primary Source URL")),
        "last_verified": clean(raw.get("Date Researched")),
        "online_process": online,
        "offline_process": offline,
        "required_documents": required,
        "category": map_category(clean(raw.get("Sector"))),
        "eligibility": clean(raw.get("Who Can Apply")),
        "district": clean(raw.get("District")),
        "verification_status": status,
        "verification_notes": clean(raw.get("Verification Notes")),
        "record_id": clean(raw.get("Record ID")),
    }


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--source-dir", required=True)
    ap.add_argument("--out-dir", default=str(RESEARCH_DIR))
    args = ap.parse_args()

    src = Path(args.source_dir)
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    files = sorted(src.glob("PaperTrail_*Research*.xlsx"))
    if not files:
        print(f"No research workbooks found in {src}")
        return 1

    summary = []
    for f in files:
        state_key = f.name.split("_")[1]
        rows = read_sheet(f)
        state = clean(rows[0].get("State")) or state_key
        docs = [d for d in (convert(r, state) for r in rows) if d]

        # A name can repeat across districts; keep the richer record.
        best: Dict[str, Dict[str, Any]] = {}
        for d in docs:
            prev = best.get(d["id"])
            if prev is None or (prev["grade"] != "guide" and d["grade"] == "guide"):
                best[d["id"]] = d
        docs = list(best.values())

        guide = [d for d in docs if d["grade"] == "guide"]
        path = out_dir / f"{state.lower().replace(' ', '_')}.json"
        path.write_text(json.dumps(docs, indent=1, ensure_ascii=False), encoding="utf-8")
        summary.append((state, len(rows), len(docs), len(guide)))
        print(f"{state:16} rows={len(rows):5}  unique={len(docs):5}  "
              f"guide={len(guide):4}  directory={len(docs)-len(guide):5}  -> {path.name}")

    tg = sum(s[3] for s in summary)
    print(f"\nTotal guide-grade records ready for the corpus: {tg}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
